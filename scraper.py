def excel_file_scraper(excel_file) -> list:
    """Returns list with dicts per row in scrapped excel file"""
    from openpyxl import load_workbook
    wb = load_workbook(excel_file)
    worksheet = wb.active
    sheet_list = [{cell.coordinate: cell.value for cell in row_cells} for row_cells in worksheet.iter_rows()]
    return sheet_list
